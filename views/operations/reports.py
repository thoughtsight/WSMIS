from views.shared import *
from views.components.kpi_engine import KPIEngine
from views.components.chart_engine import ChartEngine




from config.settings import LABOUR_DISC_BENCH

# Import shared UI helpers from app
from ui.traffic import yoy_badge, traffic_light, tgt_badge

def render(df, pairs, comparison_mode=True, selected_months=None):
    with st.spinner("Computing Reports..."):
        if df.empty:
            EmptyState('No data available for the selected period. Adjust your filters or check data freshness.')
            return
    pp_months = [p[1] for p in pairs]
    # df is already filtered by selected_months at main level, use it directly for current period
    cp = df
    pp = apply_month_filter(df, "Month Name", pp_months)
    
    st.markdown('<div class="section-card"><div class="section-title">📄 Report Generator</div>', unsafe_allow_html=True)
    
    # Report 1: Monthly MIS Summary (Excel)
    st.markdown("### 1. Monthly MIS Summary (Excel)")
    if st.button("Generate Excel Report", key="report1_btn"):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Sheet 1: Location summary
                loc_summ = location_summary(cp, as_index=True).agg(
                    JCs=("JC_Nos.","sum"),
                    NL=("Net_Labour","sum"),
                    NP=("Net_Parts","sum"),
                    M=("Total Margin","sum"),
                    DL=("Labour Discount","sum"),
                    PL=("Pre-GST Labour","sum")
                ).reset_index()
                loc_summ["Disc%"] = np.where(loc_summ["PL"]>0, loc_summ["DL"]/loc_summ["PL"]*100, 0)
                loc_summ.to_excel(writer, sheet_name='Location Summary', index=False)
                
                # Sheet 2: Advisor summary
                adv_summ = advisor_summary(cp, adv_col=ADV_COL, as_index=True).agg(
                    JCs=("JC_Nos.","sum"),
                    NL=("Net_Labour","sum"),
                    NP=("Net_Parts","sum"),
                    DL=("Labour Discount","sum"),
                    PL=("Pre-GST Labour","sum")
                ).reset_index()
                adv_summ["Disc%"] = np.where(adv_summ["PL"]>0, adv_summ["DL"]/adv_summ["PL"]*100, 0)
                adv_summ.to_excel(writer, sheet_name='Advisor Summary', index=False)
                
                # Sheet 3: Raw data
                cp.to_excel(writer, sheet_name='Raw Data', index=False)
            
            output.seek(0)
            cp_months_list = sorted(cp["Month Name"].unique(), key=lambda x: MONTH_SORT_ORDER.get(x, 99))
            filename_start = cp_months_list[0] if cp_months_list else "start"
            filename_end = cp_months_list[-1] if cp_months_list else "end"
            st.download_button(
                "📥 Download MIS Report",
                output.getvalue(),
                file_name=f"MIS_Report_{filename_start}_{filename_end}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="report1_download"
            )
        except ImportError:
            st.error("openpyxl not installed")
    
    st.markdown("---")
    
    # Report 2: Advisor Performance Report
    st.markdown("### 2. Advisor Performance Report")
    if st.button("Generate Advisor Report", key="report2_btn"):
        try:
            import openpyxl
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                aa = advisor_summary(cp, adv_col=ADV_COL, as_index=True).agg(
                    JCs=("JC_Nos.","sum"),
                    NL=("Net_Labour","sum"),
                    NP=("Net_Parts","sum"),
                    DL=("Labour Discount","sum"),
                    PL=("Pre-GST Labour","sum"),
                    Loc=("Location Name", "first")
                ).reset_index()
                aa["Avg_Lab_JC"] = np.where(aa["JCs"]>0, aa["NL"]/aa["JCs"], 0)
                aa["Disc%"] = np.where(aa["PL"]>0, aa["DL"]/aa["PL"]*100, 0)
                aa = aa.sort_values("Avg_Lab_JC", ascending=False)
                aa.to_excel(writer, sheet_name='Advisor Performance', index=False)
            
            output.seek(0)
            st.download_button(
                "📥 Download Advisor Report",
                output.getvalue(),
                file_name=f"Advisor_Performance_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="report2_download"
            )
        except ImportError:
            st.error("openpyxl not installed")
    
    st.markdown("---")
    
    # Report 3: Discount Leakage Report
    st.markdown("### 3. Discount Leakage Report")
    if st.button("Generate Leakage Report", key="report3_btn"):
        try:
            import openpyxl
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                leak = cp.groupby([ADV_COL, "Location Name"], dropna=False).agg(
                    GL=("Pre-GST Labour","sum"),
                    DL=("Labour Discount","sum")
                ).reset_index()
                leak["Disc%"] = np.where(leak["GL"]>0, leak["DL"]/leak["GL"]*100, 0)
                leak["Revenue_Lost"] = np.maximum(0, leak["DL"] - (leak["GL"] * LABOUR_DISC_BENCH / 100))
                leak = leak.sort_values("Revenue_Lost", ascending=False)
                leak.to_excel(writer, sheet_name='Discount Leakage', index=False)
                
                # Total row
                total = pd.DataFrame([{
                    ADV_COL: "TOTAL",
                    "Location Name": "",
                    "GL": leak["GL"].sum(),
                    "DL": leak["DL"].sum(),
                    "Disc%": calc_ratio(leak["DL"].sum(), leak["GL"].sum(), multiplier=100, fill_value=0),
                    "Revenue_Lost": leak["Revenue_Lost"].sum()
                }])
                total.to_excel(writer, sheet_name='Discount Leakage', index=False, startrow=len(leak)+2, header=False)
            
            output.seek(0)
            st.download_button(
                "📥 Download Leakage Report",
                output.getvalue(),
                file_name=f"Discount_Leakage_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="report3_download"
            )
        except ImportError:
            st.error("openpyxl not installed")

    st.markdown("---")

    # Report 4: Advisor Incentive Sheet
    st.markdown("### 4. Advisor Incentive Sheet")
    if st.button("Generate Incentive Sheet", key="report4_btn"):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Calculate incentive metrics
                aa = advisor_summary(cp, adv_col=ADV_COL, as_index=True).agg(
                    JCs=("JC_Nos.","sum"),
                    NL=("Net_Labour","sum"),
                    NP=("Net_Parts","sum"),
                    DL=("Labour Discount","sum"),
                    PL=("Pre-GST Labour","sum"),
                    Loc=("Location Name", lambda x: ", ".join(sorted(x.unique())) if len(x.unique()) > 1 else x.iloc[0])
                ).reset_index()
                aa["Avg_Lab_JC"] = np.where(aa["JCs"]>0, aa["NL"]/aa["JCs"], 0)
                aa["Disc%"] = np.where(aa["PL"]>0, aa["DL"]/aa["PL"]*100, 0)
                aa["Incentive"] = aa["NL"] * 0.01  # 1% of Net Labour as incentive
                aa = aa.sort_values("Incentive", ascending=False)

                # Write to Excel with formatting
                aa.to_excel(writer, sheet_name='Incentive Sheet', index=False)
                worksheet = writer.sheets['Incentive Sheet']

                # Format header
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            output.seek(0)
            st.download_button(
                "📥 Download Incentive Sheet",
                output.getvalue(),
                file_name=f"Advisor_Incentive_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="report4_download"
            )
        except ImportError:
            st.error("openpyxl not installed")

    st.markdown("---")

    # Report 5: AI Narrative Summary
    st.markdown("### 5. AI Narrative Summary")
    
    def generate_narrative(df_cp, df_pp, pairs):
        cp_jc = get_jobcard_count(df_cp)
        pp_jc = get_jobcard_count(df_pp)
        jc_growth = calc_growth_pct(cp_jc, pp_jc, fill_value=0)
        
        cp_lab = get_net_labour(df_cp)
        pp_lab = get_net_labour(df_pp)
        lab_growth = calc_growth_pct(cp_lab, pp_lab, fill_value=0)
        
        cp_margin = get_total_margin(df_cp)
        
        loc_perf = location_summary(df_cp, as_index=True)['Net_Labour'].sum()
        best_loc = loc_perf.idxmax() if not getattr(locals().get('cp', None), 'empty', True) else "N/A" if not loc_perf.empty and loc_perf.max() > 0 else "N/A"
        worst_loc = loc_perf.idxmin() if not loc_perf.empty else "N/A"

        adv_disc = advisor_summary(df_cp, adv_col=ADV_COL, as_index=True).agg(L=('Pre-GST Labour','sum'), D=('Labour Discount','sum'))
        adv_disc['D%'] = calc_ratio(adv_disc['D'], adv_disc['L'], multiplier=100, fill_value=np.nan)
        _high_adv = adv_disc[adv_disc['L'] > 50000]['D%'] if len(adv_disc[adv_disc['L'] > 50000]) > 0 else adv_disc['D%']
        worst_adv = _high_adv.idxmax() if not getattr(locals().get('cp', None), 'empty', True) else "N/A" if not _high_adv.empty else "N/A"
        worst_adv_pct = adv_disc.loc[worst_adv, 'D%'] if worst_adv != "N/A" and worst_adv in adv_disc.index else 0
        
        period_label = f"{pairs[0][0]} to {pairs[-1][0]}" if pairs else "selected period"
        
        narrative = f"""
WORKSHOP PERFORMANCE SUMMARY — {period_label}

Overall Performance: Total {int(cp_jc):,} Job Cards processed, 
{"up" if jc_growth >= 0 else "down"} {abs(jc_growth):.1f}% vs same period last year.
Net Labour revenue: {fmt_inr(cp_lab)} ({("+" if lab_growth >= 0 else "")}{lab_growth:.1f}% YoY).
Total Margin earned: {fmt_inr(cp_margin)}.

Location Highlights: {best_loc} led all locations in net labour.
{worst_loc} recorded the lowest revenue and may need attention.

Discount Alert: Advisor {worst_adv} had the highest labour discount 
at {worst_adv_pct:.1f}% — review recommended.

VOR Charges: {fmt_inr(get_vor_charges(df_cp))} — 
{"within acceptable range." if get_vor_charges(df_cp) > -100000 else "elevated — management review needed."}
        """.strip()
        
        return narrative
    
    if pairs:
        narrative = generate_narrative(cp, pp, pairs)
        st.text_area("Management Summary", narrative, height=200, key="narrative_text")
        
        st.download_button(
            "📥 Download Summary (.txt)",
            narrative.encode(),
            file_name=f"narrative_{datetime.now().strftime('%Y%m%d_%H%M')}.txt",
            mime="text/plain",
            key="narrative_download"
        )
    
    st.markdown('</div>', unsafe_allow_html=True)

