"""
Centralized export framework for WSMIS.

Provides CSV, Excel, and PDF export with consistent metadata,
Indian number formatting, and standardized filename conventions.

Filename convention: <Location>_<ReportName>_<YYYY-MM-DD>.<ext>
    Examples:
        Arena_Margin_2026-06-17.xlsx
        ExecutiveDashboard_2026-06-17.pdf
        All_Cockpit_2026-06-17.csv

Usage:
    from services.export_service import ExportMeta, export_csv, export_excel, export_pdf, build_filename

    meta = ExportMeta(
        report_title="Margin Analysis",
        location="Arena",
        date_range="Apr-25 – Jun-25",
    )
    data = export_excel(df, meta)
    filename = build_filename("Arena", "Margin", "xlsx")
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Set

import numpy as np
import pandas as pd

from config.settings import APP_NAME, VERSION
from services.logging_service import log_performance


# ─────────────────────────────────────────────────────────────────────────────
# Export Metadata
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ExportMeta:
    """
    Metadata embedded in every export.

    Every export automatically includes:
    - Application name  (from config.settings.APP_NAME)
    - Version           (from config.settings.VERSION)
    - Generation timestamp
    - Report title
    - Selected location filter
    - Selected date range
    """
    report_title: str
    location: str = "All Locations"
    date_range: str = ""
    app_name: str = field(default_factory=lambda: APP_NAME)
    version: str = field(default_factory=lambda: f"v{VERSION}")
    generated_at: str = field(
        default_factory=lambda: datetime.now().strftime("%d %b %Y %I:%M %p")
    )

    def as_dict(self) -> Dict[str, str]:
        """Ordered dict of all metadata fields for embedding in exports."""
        return {
            "Application": self.app_name,
            "Version": self.version,
            "Report": self.report_title,
            "Location": self.location,
            "Date Range": self.date_range or "—",
            "Generated": self.generated_at,
        }

    def filters_summary(self) -> str:
        """One-line summary of applied filters for PDF/Excel headers."""
        parts = []
        if self.location and self.location != "All Locations":
            parts.append(f"Location: {self.location}")
        if self.date_range:
            parts.append(f"Period: {self.date_range}")
        return "  •  ".join(parts) if parts else "No filters applied"


# ─────────────────────────────────────────────────────────────────────────────
# Filename Builder
# ─────────────────────────────────────────────────────────────────────────────

def build_filename(location: str, report_name: str, ext: str) -> str:
    """
    Construct a standardized export filename.

    Pattern: <Location>_<ReportName>_<YYYY-MM-DD>.<ext>

    Args:
        location:    Location label (e.g. "Arena", "All").
        report_name: Report identifier (e.g. "Margin", "ExecutiveDashboard").
        ext:         File extension without dot (e.g. "csv", "xlsx", "pdf").

    Returns:
        Filename string, e.g. "Arena_Margin_2026-06-17.xlsx"
    """
    date_str = datetime.now().strftime("%Y-%m-%d")
    loc = re.sub(r"[^A-Za-z0-9]", "", location) or "All"
    rpt = re.sub(r"[^A-Za-z0-9]", "", report_name) or "Report"
    return f"{loc}_{rpt}_{date_str}.{ext}"


# ─────────────────────────────────────────────────────────────────────────────
# Indian Number Formatting Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _fmt_inr(v) -> str:
    """Indian rupee format: ₹1,00,00,000 = 1 Crore, ₹1,00,000 = 1 Lakh."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    v = float(v)
    if v == 0:
        return "—"
    neg = v < 0
    a = abs(v)
    if a >= 1e7:
        s = f"₹{a / 1e7:,.2f} Cr"
    elif a >= 1e5:
        s = f"₹{a / 1e5:,.2f} L"
    elif a >= 1e3:
        s = f"₹{a / 1e3:,.1f} K"
    else:
        s = f"₹{a:,.0f}"
    return f"-{s}" if neg else s


def _fmt_pct(v) -> str:
    """Percentage with 2 decimal places."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    return f"{float(v):.2f}%"


def _detect_currency_cols(df: pd.DataFrame) -> Set[str]:
    """Heuristically identify INR amount columns."""
    keywords = {
        "sales", "revenue", "margin", "discount", "profit", "labour",
        "parts", "income", "charges", "amount", "value", "cost",
        "net", "gross", "total", "oil", "tyre", "battery", "accessory",
        "fsc", "otc", "vor",
    }
    cols: Set[str] = set()
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            col_lower = col.lower().replace("_", " ")
            if any(kw in col_lower for kw in keywords):
                cols.add(col)
    return cols


def _detect_pct_cols(df: pd.DataFrame) -> Set[str]:
    """Heuristically identify percentage columns."""
    cols: Set[str] = set()
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            col_lower = col.lower()
            if col_lower.endswith("%") or col_lower.endswith("pct") or "percent" in col_lower:
                cols.add(col)
    return cols


# ─────────────────────────────────────────────────────────────────────────────
# CSV Export
# ─────────────────────────────────────────────────────────────────────────────

@log_performance(page_context="Export CSV")
def export_csv(df: pd.DataFrame, meta: ExportMeta) -> bytes:
    """
    Export DataFrame to UTF-8 CSV with metadata comment header.

    Returns:
        bytes ready for st.download_button.
    """
    buf = io.StringIO()

    # Comment-prefixed metadata header
    for k, v in meta.as_dict().items():
        buf.write(f"# {k}: {v}\n")
    buf.write("#\n")

    df.to_csv(buf, index=False, encoding="utf-8")
    return buf.getvalue().encode("utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# Excel Export
# ─────────────────────────────────────────────────────────────────────────────

@log_performance(page_context="Export Excel")
def export_excel(df: pd.DataFrame, meta: ExportMeta) -> bytes:
    """
    Export DataFrame to a professionally formatted Excel workbook.

    Formatting applied:
    - Metadata block at top (app name, version, filters, timestamp)
    - Bold blue headers with white text
    - Frozen header row
    - Alternating row shading
    - Auto-fitted column widths
    - Indian number formatting for currency columns
    - Percentage formatting for percentage columns

    Returns:
        bytes ready for st.download_button.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. It is listed in requirements.txt.")

    # Style definitions
    BRAND_BLUE  = "0071E3"
    LIGHT_GRAY  = "F5F5F7"
    ALT_GRAY    = "F9F9FB"
    DARK_TEXT   = "1D1D1F"
    SUBTLE_TEXT = "6E6E73"
    BORDER_CLR  = "E5E5EA"

    header_fill  = PatternFill("solid", fgColor=BRAND_BLUE)
    meta_fill    = PatternFill("solid", fgColor=LIGHT_GRAY)
    alt_fill     = PatternFill("solid", fgColor=ALT_GRAY)
    thin_border  = Border(bottom=Side(style="thin", color=BORDER_CLR))

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    title_font   = Font(name="Calibri", bold=True, size=14, color=DARK_TEXT)
    meta_lbl_fnt = Font(name="Calibri", size=9,  color=SUBTLE_TEXT)
    meta_val_fnt = Font(name="Calibri", bold=True, size=9, color=DARK_TEXT)
    data_font    = Font(name="Calibri", size=10, color=DARK_TEXT)
    note_font    = Font(name="Calibri", size=8, italic=True, color=SUBTLE_TEXT)

    currency_cols = _detect_currency_cols(df)
    pct_cols      = _detect_pct_cols(df)
    num_data_cols = max(len(df.columns), 4)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = meta.report_title[:31]   # Excel sheet name limit

    # ── Row 1: Report title ──────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_data_cols)
    tc = ws.cell(row=1, column=1, value=meta.report_title)
    tc.font = title_font
    tc.fill = meta_fill
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Rows 2+: Metadata block ──────────────────────────────
    meta_items = list(meta.as_dict().items())
    for i, (k, v) in enumerate(meta_items):
        r = i + 2
        lc = ws.cell(row=r, column=1, value=k)
        lc.font = meta_lbl_fnt
        lc.fill = meta_fill
        lc.alignment = Alignment(horizontal="right", vertical="center")

        vc = ws.cell(row=r, column=2, value=v)
        vc.font = meta_val_fnt
        vc.fill = meta_fill
        vc.alignment = Alignment(horizontal="left", vertical="center")

        # Span remaining columns with meta fill
        for col_idx in range(3, num_data_cols + 1):
            blank = ws.cell(row=r, column=col_idx)
            blank.fill = meta_fill
        ws.row_dimensions[r].height = 15

    # ── Separator row ────────────────────────────────────────
    sep_row = len(meta_items) + 2
    for col_idx in range(1, num_data_cols + 1):
        ws.cell(row=sep_row, column=col_idx).fill = meta_fill
    ws.row_dimensions[sep_row].height = 8

    # ── Header row ────────────────────────────────────────────
    header_row = sep_row + 1
    for col_idx, col_name in enumerate(df.columns, start=1):
        hc = ws.cell(row=header_row, column=col_idx, value=col_name)
        hc.font = header_font
        hc.fill = header_fill
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hc.border = thin_border
    ws.row_dimensions[header_row].height = 22

    # Freeze panes below header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Data rows ────────────────────────────────────────────
    for row_offset, (_, row_data) in enumerate(df.iterrows(), start=1):
        r = header_row + row_offset
        use_alt = (row_offset % 2 == 0)

        for col_idx, col_name in enumerate(df.columns, start=1):
            raw = row_data[col_name]
            dc = ws.cell(row=r, column=col_idx)

            if col_name in currency_cols and raw is not None and not (isinstance(raw, float) and np.isnan(raw)):
                dc.value = _fmt_inr(raw)      # Indian formatted string
                dc.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in pct_cols and raw is not None and not (isinstance(raw, float) and np.isnan(raw)):
                dc.value = _fmt_pct(raw)
                dc.alignment = Alignment(horizontal="right", vertical="center")
            elif pd.api.types.is_numeric_dtype(type(raw)) and raw is not None and not (isinstance(raw, float) and np.isnan(raw)):
                dc.value = raw
                dc.alignment = Alignment(horizontal="right", vertical="center")
            else:
                dc.value = raw if (raw is not None and not (isinstance(raw, float) and np.isnan(raw))) else None
                dc.alignment = Alignment(horizontal="left", vertical="center")

            dc.font = data_font
            dc.border = thin_border
            if use_alt:
                dc.fill = alt_fill
        ws.row_dimensions[r].height = 16

    # ── Auto column widths ───────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        sample_vals = df[col_name].astype(str) if not df.empty else []
        max_val_len = sample_vals.str.len().max() if len(sample_vals) > 0 else 0
        width = min(max(len(str(col_name)), int(max_val_len or 0), 10) + 3, 45)
        ws.column_dimensions[letter].width = width

    # Ensure meta label/value columns are readable
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 20)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 32)

    # ── Footer note ─────────────────────────────────────────
    footer_row = header_row + len(df) + 2
    fc = ws.cell(row=footer_row, column=1,
                 value=f"{meta.app_name}  •  {meta.version}  •  {meta.generated_at}")
    fc.font = note_font
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=min(4, num_data_cols))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# PDF Export
# ─────────────────────────────────────────────────────────────────────────────

_PDF_ROW_LIMIT = 500   # Maximum rows in a single PDF export


@log_performance(page_context="Export PDF")
def export_pdf(df: pd.DataFrame, meta: ExportMeta) -> bytes:
    """
    Export DataFrame to a professional landscape PDF report.

    Layout:
    - Landscape A4, 1.5 cm margins
    - Header: report title, app/version subtitle, applied filters
    - Data table with bold blue header row, alternating row shading,
      repeated header on each page
    - Footer on every page: app name, version, timestamp, page number
    - Note appended if data was truncated to PDF_ROW_LIMIT

    Returns:
        bytes ready for st.download_button.

    Raises:
        ImportError: if reportlab is not installed.
    """
    try:
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable,
        )
    except ImportError:
        raise ImportError(
            "reportlab is required for PDF export. "
            "Install with: pip install reportlab"
        )

    PAGE_SIZE = landscape(A4)
    PAGE_W, PAGE_H = PAGE_SIZE
    MARGIN = 1.5 * cm
    CONTENT_W = PAGE_W - 2 * MARGIN

    # ── Color palette ────────────────────────────────────────
    C_BLUE    = rl_colors.HexColor("#0071E3")
    C_DARK    = rl_colors.HexColor("#1D1D1F")
    C_SUBTLE  = rl_colors.HexColor("#6E6E73")
    C_ALT     = rl_colors.HexColor("#F5F5F7")
    C_BORDER  = rl_colors.HexColor("#E5E5EA")

    # ── Typography styles ────────────────────────────────────
    title_style = ParagraphStyle(
        "RptTitle",
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=C_DARK,
        spaceAfter=3,
        leading=22,
    )
    sub_style = ParagraphStyle(
        "RptSub",
        fontName="Helvetica",
        fontSize=10,
        textColor=C_SUBTLE,
        spaceAfter=2,
        leading=14,
    )
    filter_style = ParagraphStyle(
        "RptFilter",
        fontName="Helvetica",
        fontSize=8,
        textColor=C_SUBTLE,
        spaceAfter=0,
        leading=11,
    )
    note_style = ParagraphStyle(
        "RptNote",
        fontName="Helvetica-Oblique",
        fontSize=7,
        textColor=C_SUBTLE,
        leading=10,
    )
    th_style = ParagraphStyle(
        "TH",
        fontName="Helvetica-Bold",
        fontSize=7,
        textColor=rl_colors.white,
        alignment=TA_CENTER,
        leading=9,
    )
    td_style = ParagraphStyle(
        "TD",
        fontName="Helvetica",
        fontSize=7,
        textColor=C_DARK,
        alignment=TA_LEFT,
        leading=9,
    )

    # ── Page footer callback ─────────────────────────────────
    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(C_BORDER)
        canvas.setLineWidth(0.5)
        canvas.line(MARGIN, 1.1 * cm, PAGE_W - MARGIN, 1.1 * cm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(C_SUBTLE)
        left = f"{meta.app_name}  •  {meta.version}  •  Generated: {meta.generated_at}"
        right = f"Page {canvas.getPageNumber()}"
        canvas.drawString(MARGIN, 0.65 * cm, left)
        canvas.drawRightString(PAGE_W - MARGIN, 0.65 * cm, right)
        canvas.restoreState()

    # ── Document ─────────────────────────────────────────────
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=PAGE_SIZE,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=MARGIN,
        bottomMargin=2.2 * cm,
    )

    story = []

    # Header block
    story.append(Paragraph(meta.report_title, title_style))
    story.append(Paragraph(f"{meta.app_name}  |  {meta.version}", sub_style))

    # Filters summary
    filter_parts: List[str] = []
    if meta.location and meta.location != "All Locations":
        filter_parts.append(f"<b>Location:</b> {meta.location}")
    if meta.date_range:
        filter_parts.append(f"<b>Date Range:</b> {meta.date_range}")
    filter_parts.append(f"<b>Generated:</b> {meta.generated_at}")
    story.append(Paragraph("  •  ".join(filter_parts), filter_style))
    story.append(HRFlowable(width=CONTENT_W, thickness=0.5, color=C_BORDER, spaceAfter=6))

    # ── Data table ───────────────────────────────────────────
    display_df = df.head(_PDF_ROW_LIMIT)
    col_names  = list(display_df.columns)
    n_cols     = len(col_names)

    # Column widths: proportional, minimum 1.4 cm
    col_w = max(1.4 * cm, CONTENT_W / max(n_cols, 1))
    col_widths = [col_w] * n_cols

    # Header row
    header_row = [Paragraph(str(c), th_style) for c in col_names]

    # Data rows
    def _cell(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "—"
        return str(v)

    data_rows = [
        [_cell(row_data[c]) for c in col_names]
        for _, row_data in display_df.iterrows()
    ]

    table_data = [header_row] + data_rows

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Build alternating row background styles
    row_styles = [
        # Header
        ("BACKGROUND",    (0, 0), (-1, 0),  C_BLUE),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  rl_colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  7),
        ("ALIGN",         (0, 0), (-1, 0),  "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, 0),  5),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  5),
        # Data rows
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 7),
        ("TEXTCOLOR",     (0, 1), (-1, -1), C_DARK),
        ("TOPPADDING",    (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        # Grid
        ("GRID",          (0, 0), (-1, -1), 0.25, C_BORDER),
        ("LINEBELOW",     (0, 0), (-1, 0),  1.0,  C_BLUE),
    ]

    # Alternating row backgrounds
    for row_idx in range(1, len(data_rows) + 1):
        if row_idx % 2 == 0:
            row_styles.append(("BACKGROUND", (0, row_idx), (-1, row_idx), C_ALT))

    tbl.setStyle(TableStyle(row_styles))
    story.append(tbl)

    if len(df) > _PDF_ROW_LIMIT:
        story.append(Spacer(1, 0.4 * cm))
        story.append(Paragraph(
            f"Note: PDF displays first {_PDF_ROW_LIMIT} of {len(df):,} rows. "
            f"Use Excel or CSV export for the complete dataset.",
            note_style,
        ))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    return buf.read()
